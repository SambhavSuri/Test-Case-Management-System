import json
import logging
import os
import time
from datetime import datetime
from pathlib import Path

import pytest
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from playwright.sync_api import sync_playwright

# Load .env from the same directory as conftest.py
load_dotenv(dotenv_path=Path(__file__).parent / ".env")

BASE_URL = os.environ["SITE_A_URL"]
USERNAME = os.environ["SITE_A_USERNAME"]
PASSWORD = os.environ["SITE_A_PASSWORD"]

HEADLESS = os.environ.get("HEADLESS", "true").lower() == "true"

REPORTS_DIR = Path(__file__).parent / "reports"
SNAPSHOTS_DIR = REPORTS_DIR / "snapshots"
REPORTS_DIR.mkdir(exist_ok=True)
SNAPSHOTS_DIR.mkdir(exist_ok=True)

# ── Logging setup ─────────────────────────────────────────────────────────────
# Each xdist worker gets its own log file to avoid garbled output.
_worker_id = os.environ.get("PYTEST_XDIST_WORKER", "")
LOG_FILE = REPORTS_DIR / (f"test_run_{_worker_id}.log" if _worker_id else "test_run.log")
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)-5s | %(message)s",
    datefmt="%H:%M:%S",
    handlers=[
        logging.FileHandler(LOG_FILE, mode="w"),
        logging.StreamHandler(),
    ],
)
logger = logging.getLogger("ui_tests")

# ── Dismiss popups (Release Notes, announcements, etc.) ───────────────────────

def dismiss_popups(page) -> None:
    """Close any overlay popups that block the UI (e.g. Release Notes modal, backdrop overlays)."""
    try:
        page.evaluate("""() => {
            // 1. Remove blocking backdrop overlays (fixed inset-0 with data-state="open")
            const backdrops = document.querySelectorAll(
                'div[data-state="open"][class*="fixed"][class*="inset-0"],'
                + 'div[data-state="open"][aria-hidden="true"][class*="backdrop"],'
                + 'div[class*="fixed"][class*="z-"][class*="bg-black"]'
            );
            for (const bd of backdrops) {
                bd.remove();
            }

            // 2. Close any open dialogs/modals by clicking their close button
            const dialogs = document.querySelectorAll('[role="dialog"][data-state="open"], [class*="modal"]');
            for (const dialog of dialogs) {
                if (dialog.offsetHeight > 0) {
                    // Find close/X button
                    const xBtn = Array.from(dialog.querySelectorAll('button'))
                        .find(b => b.querySelector('svg') && b.offsetWidth < 60);
                    if (xBtn) { xBtn.click(); continue; }
                    // Set data-state to closed
                    dialog.setAttribute('data-state', 'closed');
                    dialog.style.display = 'none';
                }
            }

            // 3. Remove any remaining fixed overlays blocking pointer events
            const fixedOverlays = document.querySelectorAll('div[class*="fixed"][class*="inset"]');
            for (const overlay of fixedOverlays) {
                const style = window.getComputedStyle(overlay);
                if (style.zIndex > 50 && overlay.offsetHeight > window.innerHeight * 0.8) {
                    overlay.remove();
                }
            }
        }""")
        page.wait_for_timeout(500)
    except Exception:
        pass


# ── Shared result collector (populated by each test) ──────────────────────────
_results: list[dict] = []


def pytest_configure(config):
    config._results = _results


@pytest.hookimpl(hookwrapper=True)
def pytest_runtest_makereport(item, call):
    """Capture screenshot and failure remarks BEFORE fixture teardown.

    This hook fires after the test call phase but before fixtures are torn down,
    so the Playwright page is still alive and we get real screenshots (not blank).
    """
    outcome = yield
    report = outcome.get_result()

    if call.when != "call":
        return

    # Store the report on the item so _record_result can access it
    item._test_report = report

    # Only take screenshots on failure
    if not report.failed:
        return

    # Find the result_entry attached by _record_result
    result_entry = getattr(item, "_result_entry", None)
    if not result_entry:
        return

    tc_id = result_entry.get("test_id") or "UNKNOWN"
    tc_name = result_entry.get("test_name", "unknown")
    snapshot_name = f"{tc_id}_{tc_name}.png"
    snapshot_path = SNAPSHOTS_DIR / snapshot_name

    # Try to grab the active page (still open at this point)
    active_page = _get_active_page_from_item(item)
    if active_page:
        try:
            active_page.screenshot(path=str(snapshot_path), full_page=True)
            result_entry["snapshot"] = snapshot_name
            logger.error(
                "%s SNAPSHOT saved → snapshots/%s  (URL: %s)",
                tc_id, snapshot_name, active_page.url,
            )
        except Exception as exc:
            logger.error("%s Could not capture snapshot: %s", tc_id, exc)
    else:
        logger.error("%s No active page found — snapshot skipped", tc_id)

    # Capture failure remarks from the assertion error
    if report.longreprtext:
        lines = report.longreprtext.strip().split("\n")
        remark = ""
        for line in reversed(lines):
            stripped = line.strip()
            # Look for "E       AssertionError: TC-XXX FAIL: ..." pattern
            if stripped.startswith("E") and "FAIL:" in stripped:
                remark = stripped.lstrip("E").strip()
                # Remove "AssertionError: " prefix if present
                if "AssertionError:" in remark:
                    remark = remark.split("AssertionError:", 1)[1].strip()
                break
            if stripped.startswith("AssertionError:"):
                remark = stripped.replace("AssertionError: ", "")
                break
        if not remark:
            # Fallback: use the actual field from result_entry
            remark = result_entry.get("actual", "")[:200]
        result_entry["remarks"] = remark[:300]


def _get_active_page_from_item(item) -> object:
    """Retrieve the Playwright page from the test item's funcargs (live fixtures)."""
    fixture_names = (
        "page", "logged_in_page", "home_page_page", "my_dealerships_page",
        "analytics_view_page", "dealer_groups_page", "single_dealership_page",
        "profile_page_page", "recalibration_page", "admin_users_page",
        "admin_groups_page",
    )
    for name in fixture_names:
        p = item.funcargs.get(name)
        if p is not None:
            try:
                if not p.is_closed():
                    return p
            except Exception:
                pass
    return None


# ── Browser / page fixtures ───────────────────────────────────────────────────

@pytest.fixture(scope="session")
def browser():
    with sync_playwright() as pw:
        b = pw.chromium.launch(headless=HEADLESS)
        yield b
        b.close()


@pytest.fixture(scope="function")
def page(browser):
    context = browser.new_context()
    p = context.new_page()
    dismiss_popups(p)
    yield p
    context.close()


# ── Credentials fixture ───────────────────────────────────────────────────────

@pytest.fixture(scope="session")
def credentials():
    return {
        "url": BASE_URL,
        "username": USERNAME,
        "password": PASSWORD,
    }


# ── Logged-in page (lands on /select after login) ────────────────────────────

@pytest.fixture(scope="function")
def logged_in_page(browser, credentials):
    """Return a page that has already completed login and is on /select."""
    from pages.login_page import LoginPage
    context = browser.new_context()
    p = context.new_page()
    lp = LoginPage(p, credentials["url"])
    lp.navigate()
    lp.login(credentials["username"], credentials["password"])
    dismiss_popups(p)
    yield p
    context.close()


# ── Home page (Login → /select → DP View → group → /kingdom/home) ────────────

GROUP_NAME = "KEN GARFF AUTOMOTIVE GROUP"


def _login_and_select_group(browser, credentials, max_retries=2):
    """Shared helper: login → DP View → select group → wait for /kingdom/home.

    Retries the full flow on transient timeouts (server throttling under load).
    Returns (context, page) on success. Raises on persistent failure.
    """
    from pages.login_page import LoginPage
    from pages.select_view_page import SelectViewPage

    last_error = None
    for attempt in range(1, max_retries + 1):
        context = browser.new_context(viewport={"width": 1920, "height": 1080})
        p = context.new_page()
        try:
            lp = LoginPage(p, credentials["url"])
            lp.navigate()
            lp.login(credentials["username"], credentials["password"])
            dismiss_popups(p)

            svp = SelectViewPage(p)
            svp.click_dp_view()
            svp.search_and_select_group(GROUP_NAME)

            p.wait_for_function(
                "() => window.location.href.includes('/kingdom/home')",
                timeout=20000,
            )
            dismiss_popups(p)
            return context, p  # success
        except Exception as exc:
            last_error = exc
            logger.warning(
                "Fixture setup attempt %d/%d failed: %s — retrying",
                attempt, max_retries, str(exc)[:100],
            )
            try:
                context.close()
            except Exception:
                pass

    # All retries exhausted — raise so pytest marks the test as ERROR (not silent skip)
    raise last_error


@pytest.fixture(scope="function")
def home_page_page(browser, credentials):
    """Return a page that has navigated to the Kingdom Home Page via DP View."""
    context, p = _login_and_select_group(browser, credentials)
    yield p
    context.close()


# ── My Dealerships page (Home → click "My Dealerships" nav) ──────────────────

@pytest.fixture(scope="function")
def my_dealerships_page(browser, credentials):
    """Return a page on the My Dealerships page (/kingdom/view?tab=dealerships)."""
    context, p = _login_and_select_group(browser, credentials)

    # Click "My Dealerships" nav button
    p.get_by_text("My Dealerships", exact=True).first.click()
    p.wait_for_load_state("domcontentloaded")
    try:
        p.wait_for_function(
            "() => window.location.href.includes('/kingdom/view')",
            timeout=15000,
        )
    except Exception:
        pass
    p.wait_for_timeout(2000)

    dismiss_popups(p)
    yield p
    context.close()


# ── Analytics View page (My Dealerships → Analysis tab) ───────────────────────

@pytest.fixture(scope="function")
def analytics_view_page(browser, credentials):
    """Return a page on the Analytics View within My Dealerships."""
    context, p = _login_and_select_group(browser, credentials)

    p.get_by_text("My Dealerships", exact=True).first.click()
    try:
        p.wait_for_function(
            "() => window.location.href.includes('/kingdom/view')",
            timeout=15000,
        )
    except Exception:
        pass
    p.wait_for_timeout(2000)

    # Click Analysis tab
    try:
        loc = p.locator('button[role="tab"]:has-text("Analysis"):visible').first
        loc.click()
    except Exception:
        try:
            p.locator('button[role="tab"]').filter(has_text="Analysis").last.click()
        except Exception:
            pass
    p.wait_for_timeout(3000)

    dismiss_popups(p)
    yield p
    context.close()


# ── Single Dealership page (GM View → group → dealer → /dashboard/home) ───────

DEALER_NAME = "KEN GARFF CHEVROLET"


@pytest.fixture(scope="function")
def single_dealership_page(browser, credentials):
    """Return a page on the Single Dealership dashboard."""
    from pages.login_page import LoginPage
    from pages.select_view_page import SelectViewPage

    last_error = None
    for attempt in range(1, 3):  # max 2 attempts
        context = browser.new_context(viewport={"width": 1920, "height": 1080})
        p = context.new_page()
        try:
            lp = LoginPage(p, credentials["url"])
            lp.navigate()
            lp.login(credentials["username"], credentials["password"])
            dismiss_popups(p)

            svp = SelectViewPage(p)
            svp.click_gm_view()
            svp.search_and_select_group(GROUP_NAME)
            svp.search_and_select_dealer(DEALER_NAME)

            p.wait_for_function(
                "() => window.location.href.includes('/dashboard')",
                timeout=20000,
            )
            p.wait_for_timeout(3000)
            dismiss_popups(p)
            break  # success
        except Exception as exc:
            last_error = exc
            logger.warning(
                "single_dealership_page attempt %d failed: %s — retrying",
                attempt, str(exc)[:100],
            )
            try:
                context.close()
            except Exception:
                pass
    else:
        raise last_error

    yield p
    context.close()


# ── Dealer Groups page (Home → click "Groups" nav) ────────────────────────────

@pytest.fixture(scope="function")
def dealer_groups_page(browser, credentials):
    """Return a page on the Dealer Groups page (/kingdom/view?tab=groups)."""
    context, p = _login_and_select_group(browser, credentials)

    # Click "Groups" nav button
    p.get_by_text("Groups", exact=True).first.click()
    p.wait_for_load_state("domcontentloaded")
    try:
        p.wait_for_function(
            "() => window.location.href.includes('tab=groups')",
            timeout=15000,
        )
    except Exception:
        pass
    p.wait_for_timeout(3000)

    dismiss_popups(p)
    yield p
    context.close()


# ── Profile page (Login → Home → avatar dropdown → Profile → /profile) ───────

@pytest.fixture(scope="function")
def profile_page_page(browser, credentials):
    """Return a page that has navigated to the Profile page."""
    from pages.home_page import HomePage
    context, p = _login_and_select_group(browser, credentials)

    # Open user dropdown → click Profile
    hp = HomePage(p)
    hp.open_user_dropdown()
    hp.click_profile()

    try:
        p.wait_for_function(
            "() => window.location.href.includes('/profile')",
            timeout=10000,
        )
    except Exception:
        pass

    dismiss_popups(p)
    yield p
    context.close()


# ── Admin Panel — Users page (Home → avatar → Admin Panel → /admin/users) ────

@pytest.fixture(scope="function")
def admin_users_page(browser, credentials):
    """Return a page on the Admin Panel Users page."""
    from pages.home_page import HomePage
    context, p = _login_and_select_group(browser, credentials)

    # Open user dropdown → click Admin Panel
    hp = HomePage(p)
    hp.open_user_dropdown()
    p.locator('[role="menuitem"]:has-text("Admin Panel")').first.click()
    try:
        p.wait_for_function(
            "() => window.location.href.includes('/admin')",
            timeout=10000,
        )
    except Exception:
        pass
    p.wait_for_timeout(2000)
    dismiss_popups(p)

    # Click Users in sidebar
    try:
        p.locator('a[href="/admin/users"]').first.click()
        p.wait_for_function(
            "() => window.location.href.includes('/admin/users')",
            timeout=10000,
        )
    except Exception:
        pass
    p.wait_for_timeout(2000)
    dismiss_popups(p)

    yield p
    context.close()


# ── Admin Panel — Groups page (Admin → sidebar Groups → /admin/groups) ───────

@pytest.fixture(scope="function")
def admin_groups_page(browser, credentials):
    """Return a page on the Admin Panel Groups page."""
    from pages.home_page import HomePage
    context, p = _login_and_select_group(browser, credentials)

    # Navigate to Admin Panel
    hp = HomePage(p)
    hp.open_user_dropdown()
    p.locator('[role="menuitem"]:has-text("Admin Panel")').first.click()
    try:
        p.wait_for_function(
            "() => window.location.href.includes('/admin')",
            timeout=10000,
        )
    except Exception:
        pass
    p.wait_for_timeout(2000)

    # Click Groups in sidebar
    p.locator('a[href="/admin/groups"]').first.click()
    try:
        p.wait_for_function(
            "() => window.location.href.includes('/admin/groups')",
            timeout=10000,
        )
    except Exception:
        pass
    p.wait_for_timeout(2000)
    dismiss_popups(p)

    yield p
    context.close()


# ── Recalibration page (Profile → Recalibration tab) ─────────────────────────

@pytest.fixture(scope="function")
def recalibration_page(browser, credentials):
    """Return a page on the Recalibration tab of the Profile page."""
    from pages.home_page import HomePage
    context, p = _login_and_select_group(browser, credentials)

    # Open user dropdown → click Profile
    hp = HomePage(p)
    hp.open_user_dropdown()
    hp.click_profile()

    try:
        p.wait_for_function(
            "() => window.location.href.includes('/profile')",
            timeout=10000,
        )
    except Exception:
        pass
    dismiss_popups(p)

    # Click Recalibration tab
    p.locator('button[role="tab"]:has-text("Recalibration")').first.click()
    p.wait_for_timeout(2000)
    dismiss_popups(p)

    yield p
    context.close()


# ── Result recording + auto-screenshot on failure ─────────────────────────────

@pytest.fixture(autouse=True)
def _record_result(request):
    """Attach a result-recording callable to each test via request.node."""
    start = time.time()
    # Derive page name from the test file: test_login.py → Login
    module = request.node.module.__name__ if request.node.module else ""
    import re
    raw = module.replace("tests.test_", "").replace("tests.", "")
    raw = re.sub(r"^\d+_", "", raw)  # remove numeric prefix like "01_"
    page_name = raw.replace("_", " ").title() if raw else "Unknown"

    result_entry: dict = {
        "test_id": "",
        "test_name": request.node.name,
        "page": page_name,
        "status": "SKIP",
        "steps": "",
        "expected": "",
        "actual": "",
        "duration": 0.0,
        "timestamp": datetime.now().isoformat(timespec="seconds"),
        "snapshot": "",
        "remarks": "",
    }

    # Expose result_entry on the item so pytest_runtest_makereport hook can access it
    request.node._result_entry = result_entry

    def record(test_id: str, steps: str, expected: str, actual: str, status: str):
        result_entry["test_id"] = test_id
        result_entry["steps"] = steps
        result_entry["expected"] = expected
        result_entry["actual"] = actual
        result_entry["status"] = status
        logger.info(
            "%s [%s] %s — %s",
            test_id,
            status,
            request.node.name,
            "PASSED" if status == "PASS" else f"FAILED: {actual}",
        )

    request.node._record = record

    yield

    result_entry["duration"] = round(time.time() - start, 2)

    # Screenshot + remarks are now captured by pytest_runtest_makereport hook
    # (runs BEFORE fixture teardown, so the page is still alive)

    _results.append(result_entry)

    # For parallel execution (pytest-xdist): persist each result to a JSON file
    # so results can be merged into one report at the end.
    try:
        result_file = REPORTS_DIR / f"_results_{_worker_id or 'main'}.json"
        existing = []
        if result_file.exists():
            existing = json.loads(result_file.read_text())
        existing.append(result_entry)
        result_file.write_text(json.dumps(existing, default=str))
    except Exception:
        pass


# ── Excel report generation ───────────────────────────────────────────────────

def pytest_sessionfinish(session, exitstatus):
    # Merge results from ALL JSON files (works for both sequential and parallel)
    all_results = []
    result_files = sorted(REPORTS_DIR.glob("_results_*.json"))
    for rf in result_files:
        try:
            all_results.extend(json.loads(rf.read_text()))
            rf.unlink()  # clean up
        except Exception:
            pass
    # Also include in-process results (for sequential runs)
    all_results.extend(_results)
    # Deduplicate by (test_id, test_name)
    seen = set()
    merged = []
    for r in all_results:
        key = (r.get("test_id", ""), r.get("test_name", ""))
        if key not in seen:
            seen.add(key)
            merged.append(r)
    all_results = merged

    if not all_results:
        return

    # Write consolidated JSON for TCMS integration
    results_json_path = REPORTS_DIR / "results_final.json"
    try:
        results_json_path.write_text(json.dumps(all_results, default=str, indent=2))
        print(f"\n📋 JSON results saved → {results_json_path}")
    except Exception as exc:
        print(f"\n⚠️ Could not save JSON results: {exc}")

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    filepath = REPORTS_DIR / f"test_report_{timestamp}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Test Results"

    headers = [
        "Test ID", "Test Name", "Page", "Steps to Reproduce",
        "Expected Result", "Actual Result", "Status",
        "Duration (s)", "Timestamp", "Snapshot", "Remarks",
    ]

    # Header row styling — black bold text on light-blue background
    header_fill = PatternFill("solid", fgColor="B8CCE4")
    header_font = Font(bold=True, color="000000", size=12)
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 25

    # Status colours
    status_colours = {
        "PASS": "C6EFCE",
        "FAIL": "FFC7CE",
        "SKIP": "FFEB9C",
    }

    # Sort results by TC number (TC-01, TC-02, …)
    def _tc_sort_key(r):
        tc = r.get("test_id", "")
        try:
            return int(tc.split("-")[1])
        except (IndexError, ValueError):
            return 9999

    sorted_results = sorted(all_results, key=_tc_sort_key)

    for row_idx, r in enumerate(sorted_results, start=2):
        snapshot_val = r.get("snapshot", "")
        remarks_val = r.get("remarks", "")
        values = [
            r["test_id"], r["test_name"], r["page"], r["steps"],
            r["expected"], r["actual"], r["status"],
            r["duration"], r["timestamp"], snapshot_val, remarks_val,
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            if col == 7:  # Status column
                colour = status_colours.get(r["status"], "FFFFFF")
                cell.fill = PatternFill("solid", fgColor=colour)
            if col == 10 and snapshot_val:  # Snapshot column — highlight if present
                cell.font = Font(color="0000FF", underline="single")
            if col == 11 and remarks_val:  # Remarks column — red text for failures
                cell.font = Font(color="CC0000")

    # Auto-fit column widths
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    wb.save(filepath)
    print(f"\n📊 Excel report saved → {filepath}")
    logger.info("Excel report saved → %s", filepath)

    # Merge worker log files into one sorted test_run.log
    main_log = REPORTS_DIR / "test_run.log"
    worker_logs = sorted(REPORTS_DIR.glob("test_run_gw*.log"))
    if worker_logs:
        all_lines = []
        for wl in worker_logs:
            try:
                all_lines.extend(wl.read_text().strip().split("\n"))
                wl.unlink()
            except Exception:
                pass
        # Sort by timestamp (first 8 chars: HH:MM:SS)
        all_lines = [l for l in all_lines if l.strip()]
        all_lines.sort(key=lambda l: l[:8] if len(l) > 8 else "")
        main_log.write_text("\n".join(all_lines) + "\n")
